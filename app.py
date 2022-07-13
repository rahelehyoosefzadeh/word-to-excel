# -*- coding: utf-8 -*-
# Author : Raheleh Yoosefzadeh
# Date : 2022-07-13

from openpyxl import load_workbook
from docx import Document
import os

# set input & output file path
# you can create the file out of the program, or you can create it here
xls_file = "workbook.xlsx"
path = "path/to/files/"

# dimensions just for example
rows_num = 1
columns_num = 7

# load workbook.xlsx and select te active sheet
work_book = load_workbook(xls_file)
sheet = work_book.active

file_counter = 1
for file_name in os.listdir(path):
    # just to make the code a little verbose to let you know what's going on
    print(str(file_counter) + " Reading from file " + file_name)

    document = Document(file_name)
    table = document.tables[0]
    for i, row in enumerate(table.rows):
        for j, cell in enumerate(row.cells):
            # Rows and Columns index in Word Document tables start from 0
            # while rows and columns in Excel workbooks start from 1
            sheet.cell(row=i + 1, column=j + 1).value = cell.text
    file_counter += 1

# Voila! save the file
work_book.save(xls_file)

print("Done!")
